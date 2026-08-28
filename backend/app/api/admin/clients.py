"""Клиенты: общая база сети (Master DB) + агрегаты по записям текущего салона."""

import io
import uuid
from datetime import UTC, datetime, timedelta
from datetime import date as date_type
from decimal import Decimal
from typing import Annotated

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from pydantic import Field
from sqlalchemy import func, select

from app.api.admin.deps import CurrentAuthor
from app.api.schemas import ApiModel, Envelope, page_meta
from app.api.security import require_salon_access
from app.models.base import Gender
from app.models.master import Client, ClientCategory, ClientImportance
from app.models.shard import AuditAction, Record, RecordStatus
from app.services.audit import diff_fields, write_audit
from app.tenancy.deps import MasterSession, TenantSession

router = APIRouter(
    prefix="/clients", tags=["clients"], dependencies=[Depends(require_salon_access)]
)

LOST_AFTER = timedelta(days=90)


# --- Схемы и агрегаты -------------------------------------------------------


class ClientOut(ApiModel):
    id: uuid.UUID
    first_name: str
    middle_name: str | None
    last_name: str | None
    phone: str
    additional_phone: str | None
    email: str | None
    birth_date: date_type | None
    gender: Gender | None
    card_number: str | None
    source: str | None
    notes: str | None
    color: str | None
    category: ClientCategory
    importance: ClientImportance
    discount_percent: int
    no_online_booking: bool
    # Агрегаты по текущему салону
    total_visits: int = 0
    total_spent: Decimal = Decimal(0)
    first_visit: datetime | None = None
    last_visit: datetime | None = None
    segment: str = "new"


class Aggregates(dict):
    """client_id → (visits, spent, first, last)."""


async def _aggregates(tenant_session, client_ids: list[uuid.UUID] | None) -> Aggregates:
    query = select(
        Record.client_id,
        func.count().filter(Record.status == RecordStatus.COMPLETED),
        func.coalesce(
            func.sum(Record.total_amount).filter(Record.status == RecordStatus.COMPLETED),
            0,
        ),
        func.min(Record.start_at).filter(Record.status == RecordStatus.COMPLETED),
        func.max(Record.start_at).filter(Record.status == RecordStatus.COMPLETED),
    ).group_by(Record.client_id)
    if client_ids is not None:
        query = query.where(Record.client_id.in_(client_ids or [uuid.uuid4()]))
    result = Aggregates()
    for cid, visits, spent, first, last in await tenant_session.execute(query):
        result[cid] = (visits, spent, first, last)
    return result


def _segment(visits: int, last: datetime | None) -> str:
    if not visits:
        return "new"
    if last and last < datetime.now(UTC) - LOST_AFTER:
        return "lost"
    return "repeat" if visits > 1 else "new"


def _client_out(client: Client, agg: Aggregates) -> ClientOut:
    visits, spent, first, last = agg.get(client.id, (0, Decimal(0), None, None))
    out = ClientOut.model_validate(client)
    out.total_visits = visits
    out.total_spent = spent
    out.first_visit = first
    out.last_visit = last
    out.segment = _segment(visits, last)
    return out


# --- CRUD -------------------------------------------------------------------


@router.get("", response_model=Envelope[list[ClientOut]])
async def list_clients(
    master_session: MasterSession,
    tenant_session: TenantSession,
    segment: str | None = None,
    category: ClientCategory | None = None,
    query_text: Annotated[str | None, Query(alias="query")] = None,
    sort: str = "name",
    page: int = 1,
    per_page: Annotated[int, Query(alias="perPage", le=200)] = 50,
) -> Envelope[list[ClientOut]]:
    # База клиентов общая для всей сети — фильтра по салону нет намеренно;
    # агрегаты и сегмент считаются по записям текущего салона.
    query = select(Client).where(Client.is_active.is_(True))
    if category is not None:
        query = query.where(Client.category == category)
    if query_text:
        pattern = f"%{query_text}%"
        query = query.where(
            Client.first_name.ilike(pattern)
            | Client.last_name.ilike(pattern)
            | Client.phone.ilike(pattern)
            | Client.card_number.ilike(pattern)
        )

    clients = list(await master_session.scalars(query))
    agg = await _aggregates(tenant_session, [c.id for c in clients])
    items = [_client_out(c, agg) for c in clients]
    if segment:
        items = [i for i in items if i.segment == segment]

    sort_keys = {
        "name": lambda i: (i.first_name, i.last_name or ""),
        "lastVisit": lambda i: i.last_visit or datetime.min.replace(tzinfo=UTC),
        "totalSpent": lambda i: i.total_spent,
        "totalVisits": lambda i: i.total_visits,
    }
    items.sort(key=sort_keys.get(sort, sort_keys["name"]), reverse=sort != "name")

    total = len(items)
    start = (page - 1) * per_page
    return Envelope(data=items[start : start + per_page], meta=page_meta(page, per_page, total))


class ClientCreateIn(ApiModel):
    first_name: str = Field(min_length=1, max_length=128)
    middle_name: str | None = None
    last_name: str | None = None
    phone: str = Field(min_length=5, max_length=32)
    additional_phone: str | None = None
    email: str | None = None
    birth_date: date_type | None = None
    gender: Gender | None = None
    card_number: str | None = None
    source: str | None = None
    notes: str | None = None
    color: str | None = None
    category: ClientCategory = ClientCategory.NEW
    importance: ClientImportance = ClientImportance.MEDIUM
    discount_percent: int = Field(default=0, ge=0, le=100)
    no_online_booking: bool = False


@router.post("", response_model=Envelope[ClientOut], status_code=status.HTTP_201_CREATED)
async def create_client(
    body: ClientCreateIn,
    author: CurrentAuthor,
    master_session: MasterSession,
    tenant_session: TenantSession,
) -> Envelope[ClientOut]:
    existing = await master_session.scalar(select(Client).where(Client.phone == body.phone))
    if existing is not None:
        raise HTTPException(409, "Клиент с таким телефоном уже существует")

    client = Client(**body.model_dump(by_alias=False))
    master_session.add(client)
    await master_session.flush()

    write_audit(
        tenant_session,
        entity="client",
        entity_id=client.id,
        entity_name=client.full_name,
        action=AuditAction.CREATED,
        author_id=author.id,
        author_name=author.name,
    )
    return Envelope(data=_client_out(client, Aggregates()))


async def _get_client(master_session, client_id: uuid.UUID) -> Client:
    client = await master_session.get(Client, client_id)
    if client is None or not client.is_active:
        raise HTTPException(404, "Клиент не найден")
    return client


@router.get("/export")
async def export_clients(
    master_session: MasterSession,
    tenant_session: TenantSession,
    include_visits: Annotated[bool, Query(alias="includeVisits")] = False,
) -> StreamingResponse:
    """Excel-выгрузка клиентской базы (+ визиты в текущем салоне)."""
    from openpyxl import Workbook

    clients = list(await master_session.scalars(select(Client).where(Client.is_active.is_(True))))
    agg = await _aggregates(tenant_session, [c.id for c in clients])

    wb = Workbook()
    ws = wb.active
    ws.title = "Клиенты"
    ws.append(
        [
            "Имя",
            "Фамилия",
            "Телефон",
            "Email",
            "Категория",
            "Скидка %",
            "Визитов",
            "Потрачено",
            "Последний визит",
        ]
    )
    for c in clients:
        visits, spent, _, last = agg.get(c.id, (0, Decimal(0), None, None))
        ws.append(
            [
                c.first_name,
                c.last_name,
                c.phone,
                c.email,
                c.category.value,
                c.discount_percent,
                visits,
                float(spent),
                last.strftime("%Y-%m-%d %H:%M") if last else "",
            ]
        )

    if include_visits:
        ws2 = wb.create_sheet("Визиты")
        ws2.append(["Клиент", "Телефон", "Мастер", "Начало", "Статус", "Сумма"])
        records = await tenant_session.scalars(select(Record).order_by(Record.start_at))
        for r in records:
            ws2.append(
                [
                    r.client_name,
                    r.client_phone,
                    r.master_name,
                    r.start_at.strftime("%Y-%m-%d %H:%M"),
                    r.status.value,
                    float(r.total_amount),
                ]
            )

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="clients.xlsx"'},
    )


class ImportReportOut(ApiModel):
    created: int
    updated: int
    errors: list[str]


@router.post("/import", response_model=Envelope[ImportReportOut])
async def import_clients(
    file: UploadFile,
    author: CurrentAuthor,
    master_session: MasterSession,
    tenant_session: TenantSession,
) -> Envelope[ImportReportOut]:
    """Импорт из Excel. Колонки листа: Имя | Фамилия | Телефон | Email.

    Клиент идентифицируется телефоном: существующий — обновляется, новый —
    создаётся. Первая строка считается заголовком и пропускается.
    """
    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(await file.read()), read_only=True)
    except Exception:  # noqa: BLE001 - openpyxl raises assorted exceptions for malformed uploads
        raise HTTPException(422, "Не удалось прочитать файл — ожидается .xlsx")

    created = updated = 0
    errors: list[str] = []
    for i, row in enumerate(wb.active.iter_rows(min_row=2, values_only=True), start=2):
        first_name, last_name, phone, email = (list(row) + [None] * 4)[:4]
        if not phone or not first_name:
            errors.append(f"Строка {i}: нужны имя и телефон")
            continue
        phone = str(phone).strip()
        client = await master_session.scalar(select(Client).where(Client.phone == phone))
        if client is None:
            master_session.add(
                Client(
                    first_name=str(first_name).strip(),
                    last_name=str(last_name).strip() if last_name else None,
                    phone=phone,
                    email=str(email).strip() if email else None,
                )
            )
            created += 1
        else:
            client.first_name = str(first_name).strip()
            client.last_name = str(last_name).strip() if last_name else None
            if email:
                client.email = str(email).strip()
            updated += 1

    write_audit(
        tenant_session,
        entity="client",
        entity_id="import",
        entity_name=f"Импорт Excel: +{created} / ~{updated}",
        action=AuditAction.CREATED,
        author_id=author.id,
        author_name=author.name,
    )
    return Envelope(data=ImportReportOut(created=created, updated=updated, errors=errors))


@router.get("/{client_id}", response_model=Envelope[ClientOut])
async def get_client(
    client_id: uuid.UUID,
    master_session: MasterSession,
    tenant_session: TenantSession,
) -> Envelope[ClientOut]:
    client = await _get_client(master_session, client_id)
    agg = await _aggregates(tenant_session, [client.id])
    return Envelope(data=_client_out(client, agg))


class ClientPatchIn(ApiModel):
    first_name: str | None = None
    middle_name: str | None = None
    last_name: str | None = None
    phone: str | None = None
    additional_phone: str | None = None
    email: str | None = None
    birth_date: date_type | None = None
    gender: Gender | None = None
    card_number: str | None = None
    source: str | None = None
    notes: str | None = None
    color: str | None = None
    category: ClientCategory | None = None
    importance: ClientImportance | None = None
    discount_percent: int | None = Field(default=None, ge=0, le=100)
    no_online_booking: bool | None = None


@router.patch("/{client_id}", response_model=Envelope[ClientOut])
async def patch_client(
    client_id: uuid.UUID,
    body: ClientPatchIn,
    author: CurrentAuthor,
    master_session: MasterSession,
    tenant_session: TenantSession,
) -> Envelope[ClientOut]:
    client = await _get_client(master_session, client_id)
    updates = body.model_dump(exclude_unset=True, by_alias=False)
    changes = diff_fields(client, updates)
    for field, value in updates.items():
        setattr(client, field, value)

    if changes:
        write_audit(
            tenant_session,
            entity="client",
            entity_id=client.id,
            entity_name=client.full_name,
            action=AuditAction.UPDATED,
            author_id=author.id,
            author_name=author.name,
            details=changes,
        )
    agg = await _aggregates(tenant_session, [client.id])
    return Envelope(data=_client_out(client, agg))


@router.delete("/{client_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_client(
    client_id: uuid.UUID,
    author: CurrentAuthor,
    master_session: MasterSession,
    tenant_session: TenantSession,
) -> None:
    """Soft-delete: на клиента ссылаются шард-БД всех салонов."""
    client = await _get_client(master_session, client_id)
    client.is_active = False
    client.deleted_at = datetime.now(UTC)
    write_audit(
        tenant_session,
        entity="client",
        entity_id=client.id,
        entity_name=client.full_name,
        action=AuditAction.DELETED,
        author_id=author.id,
        author_name=author.name,
    )


class VisitOut(ApiModel):
    id: uuid.UUID
    start_at: datetime
    service_id: uuid.UUID
    service_name: str
    master_id: uuid.UUID
    master_name: str
    status: RecordStatus
    total_amount: Decimal
    internal_notes: str | None
    photos: list[str]


@router.get("/{client_id}/visits", response_model=Envelope[list[VisitOut]])
async def client_visits(
    client_id: uuid.UUID,
    tenant_session: TenantSession,
    page: int = 1,
    per_page: Annotated[int, Query(alias="perPage", le=200)] = 50,
) -> Envelope[list[VisitOut]]:
    from app.models.shard import Service

    base = (
        select(Record, Service.name)
        .join(Service, Service.id == Record.service_id)
        .where(Record.client_id == client_id)
    )
    total = await tenant_session.scalar(select(func.count()).select_from(base.subquery()))
    rows = (
        await tenant_session.execute(
            base.order_by(Record.start_at.desc()).offset((page - 1) * per_page).limit(per_page)
        )
    ).all()
    return Envelope(
        data=[
            VisitOut(
                id=r.Record.id,
                start_at=r.Record.start_at,
                service_id=r.Record.service_id,
                service_name=r.name,
                master_id=r.Record.master_id,
                master_name=r.Record.master_name,
                status=r.Record.status,
                total_amount=r.Record.total_amount,
                internal_notes=r.Record.internal_notes,
                photos=[p.url for p in r.Record.photos],
            )
            for r in rows
        ],
        meta=page_meta(page, per_page, total or 0),
    )
